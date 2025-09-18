#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import argparse
import ssl
import sys
from pathlib import Path

from ldap3 import Server, Connection, ALL, MODIFY_REPLACE, MODIFY_DELETE, Tls
from ldap3.core.exceptions import LDAPException
from openpyxl import load_workbook

# ==== MAPEAMENTO DE APELIDOS → ATRIBUTOS LDAP ===============================
# Adicione aqui traduções de cabeçalhos comuns para nomes de atributos LDAP.
ALIAS = {
    # identificação
    "login": "sAMAccountName",
    "sam": "sAMAccountName",
    "sAMAccountName": "sAMAccountName",

    # telefone / celular
    "telefone": "mobile",
    "celular": "mobile",
    "mobile": "mobile",
    "phone": "mobile",

    # cargo/título
    "cargo": "title",
    "título": "title",
    "titulo": "title",
    "title": "title",

    # exemplos extras (adapte livremente)
    "departamento": "department",
    "department": "department",
    "empresa": "company",
    "company": "company",
    "gerencia": "manager",        # normalmente precisa do DN do manager
    "descricao": "description",
    "description": "description",
    "cidade": "l",
    "estado": "st",
    "pais": "co",
    "rua": "streetAddress",
    "cep": "postalCode",
}

# ==== ARGPARSE ==============================================================
def parse_args():
    p = argparse.ArgumentParser(
        description="Atualiza atributos no AD a partir de um XLSX: 1ª coluna = login; demais colunas = atributos."
    )
    p.add_argument("--xlsx", required=True, help="Arquivo .xlsx (1ª col: login; demais: atributos)")
    p.add_argument("--server", required=True, help="Hostname/IP do Domain Controller (ex.: dc1.empresa.local)")
    p.add_argument("--base-dn", required=True, help='Base DN (ex.: "DC=empresa,DC=local")')
    p.add_argument("--user", required=True, help='Conta com permissão (ex.: "EMPRESA\\adm.svc")')
    p.add_argument("--password", required=True, help="Senha da conta")

    p.add_argument("--sheet", default=None, help="Nome da planilha (aba). Se omitido, usa a ativa.")
    p.add_argument("--skip-header", action="store_true",
                   help="Se a primeira linha é cabeçalho. (Recomendado)")

    # Segurança da conexão
    p.add_argument("--security", choices=["ldap", "starttls", "ldaps"], default="ldaps",
                   help="Tipo de conexão. Padrão: ldaps")
    p.add_argument("--insecure", action="store_true",
                   help="Não valida certificado TLS (apenas testes).")

    # Execução
    p.add_argument("--dry-run", action="store_true",
                   help="Somente simula mudanças (não escreve no AD).")
    return p.parse_args()

# ==== CONEXÃO LDAP ==========================================================
def connect(server_host, user, password, security="ldaps", insecure=False):
    tls = None
    if security in ("starttls", "ldaps"):
        validate = ssl.CERT_NONE if insecure else ssl.CERT_REQUIRED
        tls = Tls(validate=validate)

    use_ssl = (security == "ldaps")
    server = Server(server_host, use_ssl=use_ssl, get_info=ALL, tls=tls)

    try:
        conn = Connection(server, user=user, password=password, auto_bind=False)
        conn.open()
        if security == "starttls":
            conn.start_tls()
        conn.bind()
        if not conn.bound:
            raise LDAPException("Falha ao autenticar no AD (bind).")
        return conn
    except LDAPException as e:
        raise RuntimeError(f"Erro de conexão/autenticação LDAP: {e}")

# ==== XLSX ==================================================================
def read_table(xlsx_path, sheet_name=None, skip_header=True):
    wb = load_workbook(filename=xlsx_path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []

    if skip_header:
        headers_raw = [str(h or "").strip() for h in rows[0]]
        data_rows = rows[1:]
    else:
        # Sem cabeçalho: cria nomes genericos col1, col2, ...
        headers_raw = [f"col{i+1}" for i in range(len(rows[0]))]
        data_rows = rows

    return headers_raw, data_rows

def normalize_headers(headers_raw):
    """
    Converte cabeçalhos para atributos LDAP usando ALIAS.
    Retorna:
        - login_key: o nome normalizado da 1ª coluna (precisa virar sAMAccountName)
        - attr_keys: lista de nomes de atributos LDAP para as demais colunas
    """
    if not headers_raw or len(headers_raw) < 1:
        raise ValueError("Planilha sem colunas. A 1ª coluna deve ser o login.")

    # 1ª coluna precisa mapear para sAMAccountName
    first = headers_raw[0].strip()
    login_attr = ALIAS.get(first, first)
    if login_attr.lower() != "samaccountname":
        # permitimos 'login', 'sam', 'sAMAccountName' etc. via ALIAS
        raise ValueError(
            f"A 1ª coluna deve ser o login (ex.: 'login', 'sam' ou 'sAMAccountName'). Encontrado: '{first}'"
        )

    # Demais colunas: traduzir por ALIAS, mantendo nome original se não houver
    attr_keys = []
    for h in headers_raw[1:]:
        key = (h or "").strip()
        if not key:
            attr_keys.append(None)
            continue
        attr_keys.append(ALIAS.get(key, key))  # se não tiver alias, usa direto

    return "sAMAccountName", attr_keys

# ==== LDAP HELPERS ==========================================================
def find_user_dn(conn, base_dn, sam):
    if not sam:
        return None
    filt = f"(&(objectClass=user)(sAMAccountName={sam}))"
    conn.search(search_base=base_dn, search_filter=filt, attributes=["distinguishedName"])
    if not conn.entries:
        return None
    return str(conn.entries[0].entry_dn)

def build_mods(attr_values: dict):
    """
    Gera dict de modificações LDAP:
      vazio/None -> DELETE
      valor str  -> REPLACE
    """
    mods = {}
    for attr, val in attr_values.items():
        # pula atributos sem nome (coluna vazia)
        if not attr:
            continue
        if val is None or str(val).strip() == "":
            mods[attr] = [(MODIFY_DELETE, [])]
        else:
            # força string limpa; para arrays/multivalores, adapte aqui
            mods[attr] = [(MODIFY_REPLACE, [str(val).strip()])]
    return mods

# ==== MAIN ==================================================================
def main():
    args = parse_args()

    xlsx = Path(args.xlsx)
    if not xlsx.exists():
        print(f"Arquivo não encontrado: {xlsx}", file=sys.stderr)
        sys.exit(1)

    # Lê planilha
    headers_raw, data_rows = read_table(xlsx, sheet_name=args.sheet, skip_header=args.skip_header)
    if not data_rows:
        print("Nenhuma linha de dados encontrada.", file=sys.stderr)
        sys.exit(0)

    try:
        login_key, attr_keys = normalize_headers(headers_raw)
    except ValueError as e:
        print(f"Erro nos cabeçalhos do XLSX: {e}", file=sys.stderr)
        sys.exit(1)

    # Conecta no AD
    try:
        conn = connect(
            server_host=args.server,
            user=args.user,
            password=args.password,
            security=args.security,
            insecure=args.insecure,
        )
    except RuntimeError as e:
        print(str(e), file=sys.stderr)
        sys.exit(2)

    ok, not_found, errors, total = 0, 0, 0, 0

    for idx, row in enumerate(data_rows, start=(2 if args.skip_header else 1)):
        row = list(row) if row else []
        if not row:
            continue

        # Extrai login (1ª coluna)
        sam = (str(row[0]).strip() if row[0] is not None else "")
        if not sam:
            print(f"[LINHA {idx}] Sem login na 1ª coluna — pulando.")
            continue

        # Mapeia atributos (resto das colunas)
        attr_values = {}
        for col_i, attr in enumerate(attr_keys, start=2):
            val = row[col_i - 1] if col_i - 1 < len(row) else None
            # normaliza só se não for None
            if isinstance(val, str):
                vclean = val.strip()
            else:
                vclean = str(val).strip() if val is not None else None
            attr_values[attr] = vclean

        # Remove atributos None (colunas vazias sem cabeçalho)
        attr_values = {k: v for k, v in attr_values.items() if k}

        # Ignora linhas com somente login e sem nenhum atributo para mexer
        if not attr_values:
            print(f"[LINHA {idx}] {sam} - nenhuma coluna de atributo preenchida — pulando.")
            continue

        # Busca DN
        dn = find_user_dn(conn, args.base_dn, sam)
        if not dn:
            print(f"[LINHA {idx}] Usuário não encontrado (sAMAccountName={sam}).")
            not_found += 1
            continue

        # Monta modificações
        mods = build_mods(attr_values)

        # Log amigável
        preview = ", ".join([f"{k}='{(v if (v is None or (isinstance(v, str) and v.strip()=='')) else v)}'"
                             for k, v in attr_values.items()])
        print(f"[LINHA {idx}] {sam} -> DN={dn} | {preview or '(sem mudanças)'}")

        total += 1
        if args.dry_run:
            continue

        try:
            if not conn.modify(dn, mods):
                msg = conn.result or {}
                if msg.get("result") != 0:
                    print(f"   ERRO ao modificar: {msg}", file=sys.stderr)
                    errors += 1
                else:
                    ok += 1
            else:
                ok += 1
        except LDAPException as e:
            print(f"   EXCEPTION ao modificar: {e}", file=sys.stderr)
            errors += 1

    print("\nResumo:")
    print(f"  Linhas processadas (com atributos): {total}")
    print(f"  Sucesso: {ok}")
    print(f"  Usuários não encontrados: {not_found}")
    print(f"  Erros: {errors}")

    conn.unbind()

if __name__ == "__main__":
    main()
